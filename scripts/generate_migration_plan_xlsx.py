"""Generate the ADO -> GitHub migration plan Excel workbook.

Adheres to the project Excel Output Standard defined in CLAUDE.md:
- Save to Documents/ folder
- Filename: 'AWS Inventory ...' style -> here 'ADO to GitHub Migration Plan.xlsx'
- Row 1 blank, Column A blank, header at row 2 from column B
- Calibri 11, navy #002060 header white bold centered wrap, row height 28
- Body left-aligned, vertically middle-aligned, wrap-text, thin black borders
- Auto-fit column widths capped at 60 chars, TOTAL rows bold
- Sheet 1 = Summary, then detail sheets
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parent.parent
DOCUMENTS_DIR = PROJECT_ROOT / "Documents"
DOCUMENTS_DIR.mkdir(exist_ok=True)
OUTPUT = DOCUMENTS_DIR / "ADO to GitHub Migration Plan.xlsx"

NAVY = "002060"
WHITE = "FFFFFF"

thin = Side(style="thin", color="000000")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

HEADER_FONT = Font(name="Calibri", size=11, bold=True, color=WHITE)
BODY_FONT = Font(name="Calibri", size=11)
BODY_BOLD = Font(name="Calibri", size=11, bold=True)

HEADER_FILL = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")

HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
BODY_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)


def write_sheet(ws, headers, rows, total_row_indices=None):
    total_row_indices = total_row_indices or set()

    # Row 1 blank, Column A blank -> data starts at B2
    for col_idx, header in enumerate(headers, start=2):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER
    ws.row_dimensions[2].height = 28

    for r_offset, row in enumerate(rows):
        excel_row = 3 + r_offset
        is_total = r_offset in total_row_indices
        for c_offset, value in enumerate(row):
            cell = ws.cell(row=excel_row, column=2 + c_offset, value=value)
            cell.font = BODY_BOLD if is_total else BODY_FONT
            cell.alignment = BODY_ALIGN
            cell.border = BORDER

    # Auto-fit column widths capped at 60
    for col_idx in range(2, 2 + len(headers)):
        letter = get_column_letter(col_idx)
        max_len = len(str(headers[col_idx - 2]))
        for r_offset in range(len(rows)):
            val = rows[r_offset][col_idx - 2]
            if val is None:
                continue
            for line in str(val).split("\n"):
                if len(line) > max_len:
                    max_len = len(line)
        ws.column_dimensions[letter].width = min(max_len + 2, 60)


wb = Workbook()

# ---------- Sheet 1: Summary ----------
ws = wb.active
ws.title = "Summary"
summary_rows = [
    ["Phase 1 - Discovery & Inventory", 6],
    ["Phase 2 - Design & Governance", 5],
    ["Phase 3 - Pilot Migration", 5],
    ["Phase 4 - Wave-based Cutover", 6],
    ["Phase 5 - Access & Topics Application", 4],
    ["Phase 6 - Validation & Handover", 4],
    ["TOTAL", 30],
]
write_sheet(ws, ["Phase", "Task Count"], summary_rows, total_row_indices={6})

# ---------- Sheet 2: Migration Plan ----------
ws2 = wb.create_sheet("Migration Plan")
plan = [
    # Phase 1 - Discovery & Inventory
    (1, "Phase 1 - Discovery", "Extract full inventory of ADO org via REST API: projects, repos, default branches, branch policies, size, contributors, last-commit date", "Not Started"),
    (2, "Phase 1 - Discovery", "Identify repo-name collisions across ADO projects (same repo name in multiple projects). Flag for suffix decision", "Not Started"),
    (3, "Phase 1 - Discovery", "Export ADO permissions: project-level groups, team memberships, per-repo overrides", "Not Started"),
    (4, "Phase 1 - Discovery", "Snapshot each repo locally via ghorg or scripted clone as pre-migration safety net", "Not Started"),
    (5, "Phase 1 - Discovery", "Inventory current branch policies per repo (required reviewers, min approvers, build validation, status checks, path filters)", "Not Started"),
    (6, "Phase 1 - Discovery", "Publish discovery workbook (this file + Repo Inventory sheet) to stakeholders for sign-off", "Not Started"),

    # Phase 2 - Design & Governance
    (7, "Phase 2 - Design", "Confirm GHEC tenant setup: standard GHEC (not EMU), Entra SSO + SCIM configured", "Not Started"),
    (8, "Phase 2 - Design", "Define topic taxonomy: project-*, tier-*, type-*, tech-* (controlled vocabulary in a CSV)", "Not Started"),
    (9, "Phase 2 - Design", "Define team taxonomy: <project>-admins / -maintainers / -writers / -readers + platform-admins + security-readers", "Not Started"),
    (10, "Phase 2 - Design", "Author Ruleset JSON library: default-protected, release-branches, hotfix, databricks-uat. Store in github-config repo", "Not Started"),
    (11, "Phase 2 - Design", "Map ADO permissions -> GitHub repo roles (Read / Triage / Write / Maintain / Admin). Document mapping table", "Not Started"),

    # Phase 3 - Pilot Migration
    (12, "Phase 3 - Pilot", "Select 2 low-risk repos from 2 different ADO projects for pilot", "Not Started"),
    (13, "Phase 3 - Pilot", "Run gh ado2gh migrate-repo in dry-run then live for pilots; validate Git history, tags, PRs", "Not Started"),
    (14, "Phase 3 - Pilot", "Apply topics, teams, and default Ruleset to pilot repos", "Not Started"),
    (15, "Phase 3 - Pilot", "Validate access with real users from each mapped role; capture friction", "Not Started"),
    (16, "Phase 3 - Pilot", "Retro on pilot; adjust scripts, topic list, Ruleset JSON before scaling", "Not Started"),

    # Phase 4 - Wave-based Cutover
    (17, "Phase 4 - Cutover", "Group remaining repos into waves of ~15-20 by project + risk tier", "Not Started"),
    (18, "Phase 4 - Cutover", "For each wave: freeze ADO writes (deny push via ADO policy), notify stakeholders 48h prior", "Not Started"),
    (19, "Phase 4 - Cutover", "Run gh ado2gh migrate-repo in batch for the wave; monitor GEI queue", "Not Started"),
    (20, "Phase 4 - Cutover", "Verify each migrated repo: commit SHA parity, branch count, tag count, open PR count", "Not Started"),
    (21, "Phase 4 - Cutover", "Redirect any consumers (submodules, CI clone URLs, package feeds) to new GitHub URL", "Not Started"),
    (22, "Phase 4 - Cutover", "Archive the ADO repo (read-only) once wave validation is green", "Not Started"),

    # Phase 5 - Access & Topics Application
    (23, "Phase 5 - Apply", "Bulk-apply topics per repo via gh api (project-*, tier-*, type-*, tech-*) from discovery workbook", "Not Started"),
    (24, "Phase 5 - Apply", "Bulk-create teams and assign repo permissions via gh api (dry-run first)", "Not Started"),
    (25, "Phase 5 - Apply", "Attach Rulesets to repos matching topic patterns (e.g. topic:project-a -> default-protected)", "Not Started"),
    (26, "Phase 5 - Apply", "Configure CODEOWNERS files where per-path review is required", "Not Started"),

    # Phase 6 - Validation & Handover
    (27, "Phase 6 - Validation", "Run full audit: every repo has >=1 project-* topic, correct default branch, at least one Ruleset attached", "Not Started"),
    (28, "Phase 6 - Validation", "Security review: no repo left public unintentionally, secret-scanning + push-protection enabled org-wide", "Not Started"),
    (29, "Phase 6 - Validation", "Hand over runbook to platform team: adding a new repo, adding a topic, requesting access", "Not Started"),
    (30, "Phase 6 - Validation", "Sunset ADO org (retain as read-only archive for defined period, then decommission)", "Not Started"),
]
write_sheet(
    ws2,
    ["S.No", "Phase", "Description", "Status"],
    plan,
)

# ---------- Sheet 3: Tooling Shortlist ----------
ws3 = wb.create_sheet("Tooling")
tools = [
    ("GitHub Enterprise Importer (gh ado2gh)", "GitHub official", "Repos + history + PRs + wiki + work items ADO -> GHEC", "MIT / included with GHEC", "Primary migration tool"),
    ("gh CLI + gh api", "GitHub official", "Bulk topics, teams, permissions, Rulesets", "MIT", "Primary automation post-migration"),
    ("ADO REST API + Python", "Microsoft", "Extract discovery inventory from ADO side", "Free", "Discovery phase"),
    ("ghorg", "gabrie30 (community)", "Bulk clone/backup entire ADO org", "Apache-2.0", "Pre-migration safety snapshot"),
    ("git-filter-repo", "Elijah Newren", "History rewrite (strip secrets, split monorepos) before migration", "MIT", "Only if pre-migration cleanup needed"),
    ("BFG Repo-Cleaner", "rtyley", "Simpler alternative to git-filter-repo", "GPL-3.0", "Optional alternative"),
    ("Terraform integrations/github", "GitHub", "Declarative repos/teams/protection as code", "MPL-2.0", "Optional; overkill for one-time cutover"),
    ("ado2gh Ruby gem (mona-actions)", "GitHub PS community", "Batch orchestration wrappers over GEI", "MIT", "Optional convenience"),
]
write_sheet(
    ws3,
    ["Tool", "Owner", "Purpose", "License", "Recommendation"],
    tools,
)

# ---------- Sheet 4: Repo Inventory (template) ----------
ws4 = wb.create_sheet("Repo Inventory")
inv_headers = [
    "S.No",
    "ADO Project",
    "ADO Repo Name",
    "Default Branch",
    "Size (MB)",
    "Contributors",
    "Last Commit Date",
    "Current Branch Policies",
    "Target GitHub Repo Name",
    "Collision?",
    "Target Topics",
    "Target Team(s)",
    "Wave #",
    "Status",
]
inv_rows = [
    (1, "ProjectA", "payments-service", "main", "", "", "", "2 reviewers, build required", "payments-service", "No", "project-a, tier-1, service", "project-a-writers, project-a-admins", 1, "Not Started"),
    (2, "ProjectA", "shared-lib", "main", "", "", "", "1 reviewer", "shared-lib", "Check", "project-a, library", "project-a-writers", 1, "Not Started"),
    (3, "ProjectB", "shared-lib", "master", "", "", "", "no policy", "shared-lib-projectb", "Yes (rename)", "project-b, library", "project-b-writers", 2, "Not Started"),
    (4, "ProjectB", "databricks-etl", "main", "", "", "", "1 reviewer, path filter on notebooks/*", "databricks-etl", "No", "project-b, databricks, data, tier-1", "project-b-writers, data-platform", 2, "Not Started"),
]
write_sheet(ws4, inv_headers, inv_rows)

# ---------- Sheet 5: Topic Taxonomy ----------
ws5 = wb.create_sheet("Topic Taxonomy")
topics = [
    ("project-a", "Project", "Repo belongs to former ADO Project A. Exactly one project-* topic per repo (mandatory)."),
    ("project-b", "Project", "Repo belongs to former ADO Project B."),
    ("tier-1", "Tier", "Business-critical, production-serving. Strictest Ruleset."),
    ("tier-2", "Tier", "Internal / supporting. Standard Ruleset."),
    ("tier-3", "Tier", "Experimental / sandbox. Minimal protection."),
    ("service", "Type", "Deployable service."),
    ("library", "Type", "Consumed by other repos."),
    ("infrastructure", "Type", "Terraform / Bicep / IaC."),
    ("data", "Type", "Data pipelines, ETL, notebooks."),
    ("docs", "Type", "Documentation-only repo."),
    ("databricks", "Tech", "Databricks notebooks / DLT / jobs."),
    ("terraform", "Tech", "Terraform modules or stacks."),
    ("dotnet", "Tech", ".NET codebase."),
    ("python", "Tech", "Python codebase."),
    ("node", "Tech", "Node.js / TypeScript."),
]
write_sheet(ws5, ["Topic", "Category", "Meaning / Rule"], topics)

# ---------- Sheet 6: Access Model ----------
ws6 = wb.create_sheet("Access Model")
access = [
    ("<project>-admins", "Admin", "Full control incl. settings, secrets, Rulesets. Small group."),
    ("<project>-maintainers", "Maintain", "Manage issues/PRs/releases, cannot change repo settings."),
    ("<project>-writers", "Write", "Push to non-protected branches; PRs into protected."),
    ("<project>-readers", "Read", "Clone / view only."),
    ("platform-admins", "Org Owner (subset)", "Cross-cutting platform team. Owns github-config repo."),
    ("security-readers", "Security Manager", "Read + security alerts across all repos."),
]
write_sheet(ws6, ["GitHub Team", "Repo Role", "Purpose"], access)

wb.save(OUTPUT)
print(f"Workbook saved to: {OUTPUT}")
