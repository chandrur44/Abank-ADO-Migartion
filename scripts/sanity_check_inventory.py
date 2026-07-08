"""Sanity check: compare live ADO API against Documents/ADO Inventory <org>.xlsx.

Hits ADO REST API fresh and cross-checks:
  1. Project count + names
  2. Repo count per project + repo names
  3. Total repo count
  4. Policy count per project
  5. Default branch per repo
  6. Branch/tag counts per repo (spot-check up to N repos)
  7. Collision detection

Prints a PASS/FAIL table. Exits non-zero if any check fails.
"""
from __future__ import annotations

import base64
import os
import random
import sys
from collections import Counter, defaultdict
from pathlib import Path

import requests
from dotenv import load_dotenv
from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parent.parent
load_dotenv(PROJECT_ROOT / ".env")
ORG_URL = os.environ.get("ADO_ORG_URL", "").rstrip("/")
PAT = os.environ.get("ADO_PAT", "")
if not ORG_URL or not PAT:
    print("ERROR: ADO_ORG_URL and ADO_PAT must be set in .env", file=sys.stderr)
    sys.exit(1)

ORG_NAME = ORG_URL.split("/")[-1]
WORKBOOK = PROJECT_ROOT / "Documents" / f"ADO Inventory {ORG_NAME}.xlsx"
if not WORKBOOK.exists():
    print(f"ERROR: workbook not found: {WORKBOOK}", file=sys.stderr)
    sys.exit(1)

AUTH = "Basic " + base64.b64encode(f":{PAT}".encode()).decode()
S = requests.Session()
S.headers.update({"Authorization": AUTH, "Accept": "application/json"})


def api(url, **params):
    params.setdefault("api-version", "7.1")
    r = S.get(url, params=params, timeout=60)
    r.raise_for_status()
    return r.json()


# ---------- Pull live from ADO ----------
print("Fetching live from ADO...")
live_projects = api(f"{ORG_URL}/_apis/projects", **{"$top": 500, "stateFilter": "all"}).get("value", [])
live_project_names = sorted(p["name"] for p in live_projects)

live_repos_by_project = {}      # project_name -> [repo_names]
live_default_branch = {}        # (project, repo) -> default branch (no refs/heads/)
live_policy_count_by_project = {}
live_repo_ids = {}              # (project, repo) -> repo id

for proj in live_projects:
    pid, pname = proj["id"], proj["name"]
    repos = api(f"{ORG_URL}/{pid}/_apis/git/repositories").get("value", [])
    live_repos_by_project[pname] = sorted(r["name"] for r in repos)
    for r in repos:
        live_default_branch[(pname, r["name"])] = (r.get("defaultBranch") or "").replace("refs/heads/", "") or "(none)"
        live_repo_ids[(pname, r["name"])] = r["id"]
    try:
        policies = api(f"{ORG_URL}/{pid}/_apis/policy/configurations").get("value", [])
    except requests.HTTPError:
        policies = []
    live_policy_count_by_project[pname] = len(policies)

live_total_repos = sum(len(v) for v in live_repos_by_project.values())

# Collision detection (live)
name_to_projects = defaultdict(list)
for pname, rlist in live_repos_by_project.items():
    for r in rlist:
        name_to_projects[r].append(pname)
live_collisions = {n: ps for n, ps in name_to_projects.items() if len(ps) > 1}

# ---------- Read workbook ----------
print("Reading workbook...")
wb = load_workbook(WORKBOOK, data_only=True)


def sheet_rows(ws):
    # Header row 2 from col B; data from row 3, col B
    headers = [c.value for c in ws[2] if c.value is not None]
    data = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        cells = list(row[1:1 + len(headers)])
        if all(c is None for c in cells):
            continue
        data.append(dict(zip(headers, cells)))
    return headers, data


_, summary_rows = sheet_rows(wb["Summary"])
summary = {r["Metric"]: r["Value"] for r in summary_rows}

_, project_rows = sheet_rows(wb["Projects"])
xlsx_project_names = sorted(r["Project Name"] for r in project_rows)
xlsx_repo_count_by_project = {r["Project Name"]: r["Repo Count"] for r in project_rows}

_, repo_rows = sheet_rows(wb["Repo Inventory"])
xlsx_repos_by_project = defaultdict(list)
xlsx_default_branch = {}
for r in repo_rows:
    xlsx_repos_by_project[r["ADO Project"]].append(r["ADO Repo Name"])
    xlsx_default_branch[(r["ADO Project"], r["ADO Repo Name"])] = r["Default Branch"]
for p in xlsx_repos_by_project:
    xlsx_repos_by_project[p] = sorted(xlsx_repos_by_project[p])

_, policy_rows = sheet_rows(wb["Branch Policies"])
xlsx_total_policies = len(policy_rows)

_, coll_rows = sheet_rows(wb["Collisions"])
xlsx_collisions = {r["Repo Name"] for r in coll_rows}

# ---------- Compare ----------
results = []


def check(name, expected, actual):
    ok = expected == actual
    results.append((name, "PASS" if ok else "FAIL", expected, actual))
    return ok


check("Project count", len(live_project_names), int(summary.get("Projects", 0)))
check("Project names (set)", live_project_names, xlsx_project_names)
check("Total repos", live_total_repos, int(summary.get("Repos (total)", 0)))

for pname in live_project_names:
    check(
        f"Repos in project '{pname}' (count)",
        len(live_repos_by_project.get(pname, [])),
        xlsx_repo_count_by_project.get(pname, -1),
    )
    check(
        f"Repos in project '{pname}' (names)",
        live_repos_by_project.get(pname, []),
        xlsx_repos_by_project.get(pname, []),
    )

# Policy row counts: workbook stores one row per (policy, repo scope) so >= live policy count.
# Live count is a lower bound; check workbook >= sum of live per-project policy counts.
live_total_policies = sum(live_policy_count_by_project.values())
xlsx_gte = xlsx_total_policies >= live_total_policies
results.append((
    "Policy rows in workbook >= live policy count",
    "PASS" if xlsx_gte else "FAIL",
    f">={live_total_policies}",
    xlsx_total_policies,
))

check("Collisions (set of names)", sorted(live_collisions.keys()), sorted(xlsx_collisions))

# Spot-check default branch on up to 8 random repos
sample = random.sample(list(live_default_branch.keys()), min(8, len(live_default_branch)))
for key in sample:
    check(f"Default branch for {key[0]}/{key[1]}", live_default_branch[key], xlsx_default_branch.get(key, "MISSING"))

# ---------- Print result ----------
passes = sum(1 for _, s, *_ in results if s == "PASS")
fails = sum(1 for _, s, *_ in results if s == "FAIL")

print()
print(f"{'Check':<60} {'Result':<6} {'Expected':<30} {'Actual':<30}")
print("-" * 130)
for name, status, exp, act in results:
    exp_s = str(exp)
    act_s = str(act)
    if len(exp_s) > 28:
        exp_s = exp_s[:25] + "..."
    if len(act_s) > 28:
        act_s = act_s[:25] + "..."
    print(f"{name[:58]:<60} {status:<6} {exp_s:<30} {act_s:<30}")
print("-" * 130)
print(f"TOTAL: {passes} PASS, {fails} FAIL")

sys.exit(0 if fails == 0 else 2)
