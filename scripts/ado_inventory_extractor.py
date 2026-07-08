"""Azure DevOps inventory extractor.

Reads ADO_ORG_URL + ADO_PAT from .env, walks every project and repo in the
organization, and writes an Excel workbook to Documents/ following the
project's Excel Output Standard.

PAT scopes required: Code:Read, Project and Team:Read, Policy:Read.

Sheets:
  1. Summary            - org-level counts
  2. Projects           - one row per ADO project
  3. Repo Inventory     - one row per repo (metadata + target GitHub columns)
  4. Branch Policies    - one row per policy
  5. Collisions         - repo names that appear in more than one project

The token is never logged or written anywhere. .env is gitignored.
"""
from __future__ import annotations

import base64
import os
import sys
import time
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

import requests
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parent.parent
DOCUMENTS_DIR = PROJECT_ROOT / "Documents"
DOCUMENTS_DIR.mkdir(exist_ok=True)

load_dotenv(PROJECT_ROOT / ".env")
ORG_URL = os.environ.get("ADO_ORG_URL", "").rstrip("/")
PAT = os.environ.get("ADO_PAT", "")

if not ORG_URL or not PAT:
    print("ERROR: ADO_ORG_URL and ADO_PAT must be set in .env", file=sys.stderr)
    sys.exit(1)

ORG_NAME = ORG_URL.rstrip("/").split("/")[-1]
OUTPUT = DOCUMENTS_DIR / f"ADO Inventory {ORG_NAME}.xlsx"

AUTH_HEADER = "Basic " + base64.b64encode(f":{PAT}".encode()).decode()
SESSION = requests.Session()
SESSION.headers.update({"Authorization": AUTH_HEADER, "Accept": "application/json"})

API_VERSION = "7.1"


def api_get(url: str, params: dict | None = None) -> dict[str, Any]:
    params = dict(params or {})
    params.setdefault("api-version", API_VERSION)
    for attempt in range(5):
        r = SESSION.get(url, params=params, timeout=60)
        if r.status_code == 429:
            wait = int(r.headers.get("Retry-After", "10"))
            print(f"  rate-limited, sleep {wait}s")
            time.sleep(wait)
            continue
        if r.status_code == 203:
            print("ERROR: got 203 (auth redirect). PAT likely invalid or org URL wrong.", file=sys.stderr)
            sys.exit(2)
        r.raise_for_status()
        return r.json()
    raise RuntimeError(f"Failed after retries: {url}")


def fetch_projects() -> list[dict]:
    projects, top, skip = [], 200, 0
    while True:
        data = api_get(f"{ORG_URL}/_apis/projects", {"$top": top, "$skip": skip, "stateFilter": "all"})
        batch = data.get("value", [])
        projects.extend(batch)
        if len(batch) < top:
            break
        skip += top
    return projects


def fetch_repos(project_id: str) -> list[dict]:
    return api_get(f"{ORG_URL}/{project_id}/_apis/git/repositories").get("value", [])


def fetch_branches(project_id: str, repo_id: str) -> list[dict]:
    try:
        return api_get(
            f"{ORG_URL}/{project_id}/_apis/git/repositories/{repo_id}/refs",
            {"filter": "heads/"},
        ).get("value", [])
    except requests.HTTPError:
        return []


def fetch_tags(project_id: str, repo_id: str) -> list[dict]:
    try:
        return api_get(
            f"{ORG_URL}/{project_id}/_apis/git/repositories/{repo_id}/refs",
            {"filter": "tags/"},
        ).get("value", [])
    except requests.HTTPError:
        return []


def fetch_recent_commits(project_id: str, repo_id: str, top: int = 200) -> list[dict]:
    try:
        return api_get(
            f"{ORG_URL}/{project_id}/_apis/git/repositories/{repo_id}/commits",
            {"searchCriteria.$top": top},
        ).get("value", [])
    except requests.HTTPError:
        return []


def fetch_policies(project_id: str) -> list[dict]:
    try:
        return api_get(f"{ORG_URL}/{project_id}/_apis/policy/configurations").get("value", [])
    except requests.HTTPError:
        return []


# -------- Excel styling --------
NAVY = "002060"
WHITE = "FFFFFF"
thin = Side(style="thin", color="000000")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color=WHITE)
BODY_FONT = Font(name="Calibri", size=11)
BODY_BOLD = Font(name="Calibri", size=11, bold=True)
HEADER_FILL = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
BODY_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)


def write_sheet(ws, headers, rows, total_row_indices=None):
    total_row_indices = total_row_indices or set()
    for col_idx, header in enumerate(headers, start=2):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER
    ws.row_dimensions[2].height = 28
    for r_off, row in enumerate(rows):
        er = 3 + r_off
        is_total = r_off in total_row_indices
        for c_off, val in enumerate(row):
            cell = ws.cell(row=er, column=2 + c_off, value=val)
            cell.font = BODY_BOLD if is_total else BODY_FONT
            cell.alignment = BODY_ALIGN
            cell.border = BORDER
    for col_idx in range(2, 2 + len(headers)):
        letter = get_column_letter(col_idx)
        max_len = len(str(headers[col_idx - 2]))
        for r_off in range(len(rows)):
            val = rows[r_off][col_idx - 2]
            if val is None:
                continue
            for line in str(val).split("\n"):
                if len(line) > max_len:
                    max_len = len(line)
        ws.column_dimensions[letter].width = min(max_len + 2, 60)


def summarize_policy(policy: dict) -> str:
    t = policy.get("type", {}).get("displayName", "?")
    settings = policy.get("settings", {})
    bits = [t]
    if "minimumApproverCount" in settings:
        bits.append(f"min approvers={settings['minimumApproverCount']}")
    if settings.get("resetOnSourcePush"):
        bits.append("reset-on-push")
    if settings.get("requireVoteOnLastIteration"):
        bits.append("vote-on-last-iter")
    if "buildDefinitionId" in settings:
        bits.append(f"build={settings['buildDefinitionId']}")
    return "; ".join(bits)


def scope_branch(policy: dict) -> str:
    scopes = policy.get("settings", {}).get("scope", []) or []
    refs = []
    for s in scopes:
        ref = s.get("refName") or ""
        if ref.startswith("refs/heads/"):
            ref = ref[len("refs/heads/"):]
        refs.append(ref or s.get("matchKind", "?"))
    return ", ".join(refs) or "(all)"


def scope_repo_ids(policy: dict) -> set[str]:
    return {s.get("repositoryId") for s in policy.get("settings", {}).get("scope", []) if s.get("repositoryId")}


def main():
    print(f"Extracting inventory from {ORG_URL}")
    projects = fetch_projects()
    print(f"Found {len(projects)} projects")

    project_rows = []
    repo_rows = []
    policy_rows = []
    repo_name_to_projects: dict[str, list[str]] = defaultdict(list)

    project_repo_counts = {}
    total_repos = 0
    total_disabled = 0
    repos_with_policies = 0
    total_size_bytes = 0

    for pi, proj in enumerate(sorted(projects, key=lambda p: p["name"].lower()), start=1):
        pid = proj["id"]
        pname = proj["name"]
        print(f"[{pi}/{len(projects)}] Project: {pname}")

        repos = fetch_repos(pid)
        project_repo_counts[pname] = len(repos)
        total_repos += len(repos)

        policies = fetch_policies(pid)
        # Map repo_id -> [policy summaries]
        repo_policy_summaries: dict[str, list[str]] = defaultdict(list)
        for pol in policies:
            rids = scope_repo_ids(pol)
            summary = f"[{scope_branch(pol)}] {summarize_policy(pol)}" + (
                " (blocking)" if pol.get("isBlocking") else ""
            )
            targets = rids if rids else set(r["id"] for r in repos)  # project-wide
            for rid in targets:
                repo_policy_summaries[rid].append(summary)
            # Also record as a policy row
            for rid in targets:
                rname = next((r["name"] for r in repos if r["id"] == rid), "?")
                policy_rows.append((
                    len(policy_rows) + 1,
                    pname,
                    rname,
                    scope_branch(pol),
                    pol.get("type", {}).get("displayName", "?"),
                    "Yes" if pol.get("isEnabled") else "No",
                    "Yes" if pol.get("isBlocking") else "No",
                    summarize_policy(pol),
                ))

        project_rows.append((
            pi,
            pname,
            (proj.get("description") or "")[:200],
            proj.get("visibility", ""),
            len(repos),
            proj.get("lastUpdateTime", "")[:10],
            f"{ORG_URL}/{pname}",
        ))

        for repo in sorted(repos, key=lambda r: r["name"].lower()):
            rid = repo["id"]
            rname = repo["name"]
            repo_name_to_projects[rname].append(pname)

            size = repo.get("size") or 0
            total_size_bytes += size
            disabled = repo.get("isDisabled", False)
            if disabled:
                total_disabled += 1

            branches = fetch_branches(pid, rid)
            tags = fetch_tags(pid, rid)
            commits = fetch_recent_commits(pid, rid, top=200)
            unique_authors = len({c.get("author", {}).get("email", "") for c in commits if c.get("author")})
            last_commit_date = ""
            if commits:
                last_commit_date = (commits[0].get("author", {}).get("date") or commits[0].get("committer", {}).get("date") or "")[:10]

            default_branch = (repo.get("defaultBranch") or "").replace("refs/heads/", "") or "(none)"
            policies_for_repo = repo_policy_summaries.get(rid, [])
            if policies_for_repo:
                repos_with_policies += 1

            repo_rows.append((
                len(repo_rows) + 1,
                pname,
                rname,
                default_branch,
                round(size / (1024 * 1024), 2) if size else 0,
                len(branches),
                len(tags),
                unique_authors,
                last_commit_date,
                "Yes" if disabled else "No",
                repo.get("webUrl", ""),
                "\n".join(policies_for_repo) if policies_for_repo else "(none)",
                rname,  # Target GitHub Repo Name (default = same, adjust if collision)
                "",     # Collision? filled below
                f"project-{pname.lower().replace(' ', '-')}",
                "",     # Target Team(s)
                "",     # Wave #
                "Not Started",
            ))

    # Detect collisions
    collisions = {n: ps for n, ps in repo_name_to_projects.items() if len(ps) > 1}
    coll_rows = []
    coll_repo_names = set(collisions.keys())
    for i, (name, projs) in enumerate(sorted(collisions.items()), start=1):
        coll_rows.append((i, name, ", ".join(projs), f"Suffix all but one, e.g. {name}-<project>"))

    # Fill Collision? column in repo_rows
    for i, row in enumerate(repo_rows):
        rname = row[2]
        marker = "Yes" if rname in coll_repo_names else "No"
        repo_rows[i] = row[:13] + (marker,) + row[14:]

    # ---------- Build workbook ----------
    wb = Workbook()

    ws_sum = wb.active
    ws_sum.title = "Summary"
    summary_rows = [
        ("Organization", ORG_NAME),
        ("Extraction Date", datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")),
        ("Projects", len(projects)),
        ("Repos (total)", total_repos),
        ("Repos - Disabled", total_disabled),
        ("Repos - With Branch Policies", repos_with_policies),
        ("Total Repo Size (MB)", round(total_size_bytes / (1024 * 1024), 2)),
        ("Repo Name Collisions", len(collisions)),
    ]
    write_sheet(ws_sum, ["Metric", "Value"], summary_rows)

    ws_p = wb.create_sheet("Projects")
    write_sheet(
        ws_p,
        ["S.No", "Project Name", "Description", "Visibility", "Repo Count", "Last Update", "Web URL"],
        project_rows,
    )

    ws_r = wb.create_sheet("Repo Inventory")
    write_sheet(
        ws_r,
        [
            "S.No", "ADO Project", "ADO Repo Name", "Default Branch", "Size (MB)",
            "Branch Count", "Tag Count", "Contributor Count (recent)", "Last Commit Date",
            "Is Disabled", "ADO Web URL", "Current Branch Policies",
            "Target GitHub Repo Name", "Collision?", "Target Topics", "Target Team(s)",
            "Wave #", "Status",
        ],
        repo_rows,
    )

    ws_pol = wb.create_sheet("Branch Policies")
    write_sheet(
        ws_pol,
        ["S.No", "Project", "Repo", "Branch Scope", "Policy Type", "Enabled", "Blocking", "Settings"],
        policy_rows,
    )

    ws_c = wb.create_sheet("Collisions")
    write_sheet(
        ws_c,
        ["S.No", "Repo Name", "Appears In Projects", "Suggested Rename Rule"],
        coll_rows,
    )

    wb.save(OUTPUT)
    print(f"\nWorkbook saved to: {OUTPUT}")
    print(f"Projects: {len(projects)} | Repos: {total_repos} | Policies: {len(policy_rows)} | Collisions: {len(collisions)}")


if __name__ == "__main__":
    main()
