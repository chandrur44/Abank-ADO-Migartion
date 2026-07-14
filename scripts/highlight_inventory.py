"""Highlight empty repos and stale/unused projects in ADO Inventory DefaultCollection.xlsx.

Empty repos  (Default Branch == "(none)") -> yellow fill in Repo Inventory
Stale projects (only 1 repo AND last updated before 2024-01-01) -> orange fill in Projects sheet
"""
from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

WORKBOOK = (
    Path(__file__).resolve().parent.parent
    / "Abank Document"
    / "ADO Inventory DefaultCollection.xlsx"
)

EMPTY_FILL  = PatternFill(start_color="FFD7D7", end_color="FFD7D7", fill_type="solid")  # soft red
STALE_FILL  = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # yellow
NORMAL_FILL = PatternFill(fill_type=None)

STALE_CUTOFF = date(2024, 1, 1)


def col_index(ws, header_row: int, name: str) -> int | None:
    for cell in ws[header_row]:
        if cell.value and str(cell.value).strip().lower() == name.strip().lower():
            return cell.column
    return None


def highlight_repo_inventory(wb):
    ws = wb["Repo Inventory"]
    # Header is row 2, data from row 3
    hrow = 2
    col_default = col_index(ws, hrow, "Default Branch")
    if col_default is None:
        print("WARNING: 'Default Branch' column not found in Repo Inventory - skipping")
        return 0

    highlighted = 0
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        default_branch_cell = row[col_default - 1]
        val = str(default_branch_cell.value or "").strip()
        if val.lower() in ("(none)", "", "none"):
            for cell in row:
                if cell.value is not None or cell.column >= 2:
                    cell.fill = EMPTY_FILL
            highlighted += 1
    return highlighted


def highlight_projects(wb):
    ws = wb["Projects"]
    hrow = 2
    col_repo = col_index(ws, hrow, "Repo Count")
    col_last = col_index(ws, hrow, "Last Update")

    if col_repo is None or col_last is None:
        print(f"WARNING: Missing columns in Projects sheet (Repo Count={col_repo}, Last Update={col_last}) - skipping")
        return 0

    highlighted = 0
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        repo_count_cell = row[col_repo - 1]
        last_update_cell = row[col_last - 1]

        try:
            repo_count = int(repo_count_cell.value or 0)
        except (TypeError, ValueError):
            continue

        last_update_val = str(last_update_cell.value or "").strip()[:10]
        try:
            last_update = date.fromisoformat(last_update_val)
        except ValueError:
            last_update = None

        is_stale = repo_count <= 1 and (last_update is None or last_update < STALE_CUTOFF)
        if is_stale:
            for cell in row:
                if cell.value is not None or cell.column >= 2:
                    cell.fill = STALE_FILL
            highlighted += 1

    return highlighted


def main():
    if not WORKBOOK.exists():
        print(f"ERROR: Workbook not found at {WORKBOOK}")
        return

    wb = load_workbook(WORKBOOK)
    sheets = wb.sheetnames
    print(f"Opened: {WORKBOOK}")
    print(f"Sheets: {sheets}")

    if "Repo Inventory" not in sheets:
        print("ERROR: 'Repo Inventory' sheet not found")
        return
    if "Projects" not in sheets:
        print("ERROR: 'Projects' sheet not found")
        return

    n_repos = highlight_repo_inventory(wb)
    n_projects = highlight_projects(wb)

    wb.save(WORKBOOK)
    print(f"Highlighted {n_repos} empty repos (red) in Repo Inventory")
    print(f"Highlighted {n_projects} stale/unused projects (yellow) in Projects")
    print(f"Saved: {WORKBOOK}")


if __name__ == "__main__":
    main()
