"""Generate ADO -> GitHub Migration Roadmap Excel workbook.

Output: Abank Document/ADO to GitHub Migration Roadmap Abank.xlsx
Excel Output Standard per CLAUDE.md.
"""
from __future__ import annotations

from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parent.parent
ABANK_DOC_DIR = PROJECT_ROOT / "Abank Document"
ABANK_DOC_DIR.mkdir(exist_ok=True)
OUTPUT = ABANK_DOC_DIR / "ADO to GitHub Migration Roadmap Abank.xlsx"

NAVY = "002060"
WHITE = "FFFFFF"

thin = Side(style="thin", color="000000")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

HEADER_FONT = Font(name="Calibri", size=11, bold=True, color=WHITE)
BODY_FONT = Font(name="Calibri", size=11)
BODY_BOLD = Font(name="Calibri", size=11, bold=True)

HEADER_FILL = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
BODY_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)

PHASE_FILLS = {
    "Phase 0": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    "Phase 1": PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid"),
    "Phase 2": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
    "Phase 3": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
    "Phase 4": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    "Phase 5": PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid"),
    "Phase 6": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
    "Phase 7": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
}


def write_sheet(ws, headers, rows, phase_col_idx=None, bold_rows=None):
    bold_rows = bold_rows or set()
    for col_idx, header in enumerate(headers, start=2):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER
    ws.row_dimensions[2].height = 28
    for r_off, row in enumerate(rows):
        er = 3 + r_off
        phase_val = row[phase_col_idx] if phase_col_idx is not None and phase_col_idx < len(row) else None
        phase_fill = PHASE_FILLS.get(str(phase_val)) if phase_val else None
        for c_off, val in enumerate(row):
            cell = ws.cell(row=er, column=2 + c_off, value=val)
            cell.font = BODY_BOLD if r_off in bold_rows else BODY_FONT
            cell.alignment = BODY_ALIGN
            cell.border = BORDER
            if phase_fill:
                cell.fill = phase_fill
    for col_idx in range(2, 2 + len(headers)):
        letter = get_column_letter(col_idx)
        max_len = len(str(headers[col_idx - 2]))
        for r_off in range(len(rows)):
            val = rows[r_off][col_idx - 2] if col_idx - 2 < len(rows[r_off]) else ""
            if val is None:
                continue
            for line in str(val).split("\n"):
                max_len = max(max_len, len(line))
        ws.column_dimensions[letter].width = min(max_len + 2, 60)


BASE = date(2026, 7, 14)
# Project spans 43 working-day offsets; scale them to fit within 34 working days = 31 Aug 2026
_MAX_OFFSET = 43
_TARGET_DAYS = 34


def d(offset_days):
    """Return the date that is offset_days *working days* after BASE (Mon-Fri only)."""
    current = BASE
    remaining = offset_days
    while remaining > 0:
        current += timedelta(days=1)
        if current.weekday() < 5:  # 0=Mon … 4=Fri
            remaining -= 1
    return current.strftime("%d %b %Y")


def s(n):
    """Scale n to fit the full timeline within _TARGET_DAYS working days."""
    return d(round(n * _TARGET_DAYS / _MAX_OFFSET))


wb = Workbook()

# ── Sheet 1: Summary ──────────────────────────
ws_sum = wb.active
ws_sum.title = "Summary"
write_sheet(ws_sum, ["Item", "Detail"], [
    ("Client",                  "Amalgamated Bank (Abank)"),
    ("Source",                  "Azure DevOps Server on-prem (abdevopsdev, API 5.1)"),
    ("Target",                  "GitHub Enterprise Cloud (Standard)"),
    ("Total Repos",             "327"),
    ("Active Repos to Migrate", "309 (18 empty repos excluded, confirmed skip)"),
    ("ADO Projects",            "24"),
    ("Migration Approach",      "git push mirror from VDI. GEI not supported for on-prem ADO Server."),
    ("Batch Size",              "10 repos per batch"),
    ("Total Batches",           "31"),
    ("Total Phases",            "8 (Phase 0 through Phase 7)"),
    ("Planned Start",           d(0)),
    ("Planned End",             "31 Aug 2026"),
    ("Total Duration",          "Approx. 7 weeks (weekends excluded)"),
])

# ── Sheet 2: Roadmap (no Owner column) ────────
ws_road = wb.create_sheet("Roadmap")
roadmap = [
    # Phase 0
    (1,  "Phase 0", "Kickoff and Governance",
     "Align all stakeholders on scope, migration approach, and timeline. Decisions needed: master to main rename, 18 empty repos skip, 3 repo name conflicts, and PR history acceptance.",
     s(0),   s(0),   "In Progress", "Scope sign-off and decision log"),

    (2,  "Phase 0", "Inventory Finalization",
     "Confirm the 309 active repos, agree on batch groupings, resolve the 3 collision renames, and confirm disposition of the 18 empty repos.",
     s(1),   s(1),   "Not Started", "Finalized Repo Inventory with batch assignments"),

    # Phase 1 - starts 16 Jul (d+2), no gap after Phase 0
    (3,  "Phase 1", "Script - Mirror Migration",
     "Build git_mirror_migrate.py to clone each repo from ADO and push it to GitHub. Includes dry-run mode, retry on failure, resume capability, and the master to main rename option.",
     s(2),   s(6),   "Not Started", "git_mirror_migrate.py ready"),

    (4,  "Phase 1", "Script - Topics and Teams",
     "Build apply_topics_teams.py to set GitHub topics and team permissions for each repo based on the workbook. Dry-run and idempotent.",
     s(2),   s(6),   "Not Started", "apply_topics_teams.py ready"),

    (5,  "Phase 1", "Script - Branch Rulesets",
     "Build apply_rulesets.py with 3 to 4 Ruleset JSON templates covering default protection, require review, file size, and comment required. Rulesets are attached by topic.",
     s(2),   s(6),   "Not Started", "apply_rulesets.py and JSON templates ready"),

    (6,  "Phase 1", "Script - Post-Migration Verify",
     "Build post_migration_verify.py to run a pass or fail check on each repo covering branch count, default branch, topics, team access, and Ruleset attachment.",
     s(4),   s(7),   "Not Started", "post_migration_verify.py ready"),

    (7,  "Phase 1", "Script - CI/CD Template Applicator",
     "Develop the standard GitHub Actions workflow YAML for identified repos. Build apply_cicd.py to push this template to each repo automatically.",
     s(4),   s(7),   "Not Started", "cicd_template.yml and apply_cicd.py ready"),

    (8,  "Phase 1", "Self-Hosted Runner Setup",
     "Register an org-level self-hosted runner on Abank infrastructure. Validate that the runner comes online in the GitHub org and successfully picks up workflow jobs.",
     s(2),   s(7),   "Not Started", "Runner registered and online in GitHub org"),

    (9,  "Phase 1", "Script Testing on Sandbox",
     "Test all scripts against the zeb-ai sandbox org. Validate dry-run, retry, resume, and error handling for each script before using on Abank repos.",
     s(7),   s(9),   "Not Started", "All scripts passing on sandbox"),

    # Phase 2
    (10, "Phase 2", "Pilot - 2 Repos",
     "Run the full end-to-end sequence on 2 repos: one from ScriptRefactoring and one stale single-repo project. Steps are mirror, topics, Ruleset, CI/CD, and verify. Repo owner validates the result.",
     s(9),   s(10),  "Not Started", "2 repos on GitHub, verify passing"),

    (11, "Phase 2", "Pilot Review and Fix",
     "Review what gaps were found in the pilot. Fix scripts and update the runbook. Get written sign-off to proceed to batch migration.",
     s(10),  s(12),  "Not Started", "Updated scripts, runbook, and sign-off"),

    # Phase 3
    (12, "Phase 3", "Batch 1 - Repos 1 to 10",
     "Stale single-repo projects last active more than 2 years ago. Send 24 hours freeze notice, mirror, apply topics and Ruleset, verify, then archive in ADO.",
     s(12),  s(13),  "Not Started", "10 repos on GitHub, ADO archived"),

    (13, "Phase 3", "Batch 2 - Repos 11 to 20",
     "Single-repo projects last active 1 to 2 years ago. Same freeze, mirror, topics, Ruleset, verify, archive process.",
     s(13),  s(14),  "Not Started", "10 repos on GitHub, ADO archived"),

    (14, "Phase 3", "Batch 3 - Repos 21 to 30",
     "SSIS ETL repos (5) and remaining single-repo projects. Same process as Batch 1.",
     s(14),  s(15),  "Not Started", "10 repos on GitHub, ADO archived"),

    (15, "Phase 3", "Batch 4 - Repos 31 to 40",
     "Non Binary repos (2) and active single-repo projects. Send 24 hours freeze notice before starting.",
     s(15),  s(16),  "Not Started", "10 repos on GitHub, ADO archived"),

    (16, "Phase 3", "Batch 5 - Repos 41 to 50",
     "Remaining active single-repo projects. Send 24 hours freeze notice before starting.",
     s(16),  s(17),  "Not Started", "10 repos on GitHub, ADO archived"),

    # Phase 4
    (17, "Phase 4", "Batch 6 - Repos 51 to 60",
     "ScriptRefactoring project, first 10 repos. Send 24 hours freeze notice before starting.",
     s(19),  s(20),  "Not Started", "10 SR repos on GitHub"),

    (18, "Phase 4", "Batch 7 - Repos 61 to 70",
     "ScriptRefactoring project, next 10 repos.",
     s(20),  s(21),  "Not Started", "10 SR repos on GitHub"),

    (19, "Phase 4", "Batch 8 - Repos 71 to 80",
     "ScriptRefactoring project, next 10 repos.",
     s(21),  s(22),  "Not Started", "10 SR repos on GitHub"),

    (20, "Phase 4", "Batch 9 - Repos 81 to 90",
     "ScriptRefactoring project, next 10 repos.",
     s(22),  s(23),  "Not Started", "10 SR repos on GitHub"),

    (21, "Phase 4", "Batch 10 - Repos 91 to 106",
     "ScriptRefactoring project, final 16 repos. All ScriptRefactoring repos now on GitHub.",
     s(23),  s(26),  "Not Started", "All 56 SR repos on GitHub"),

    # Phase 5
    (22, "Phase 5", "Batch 11 - Repos 107 to 116",
     "AB AppDev project begins. First 10 repos ordered by oldest last commit. Send 24 hours freeze notice before starting.",
     s(26),  s(27),  "Not Started", "10 AB AppDev repos on GitHub"),

    (23, "Phase 5", "Batch 12 - Repos 117 to 126",
     "AB AppDev project, next 10 repos.",
     s(27),  s(28),  "Not Started", "10 AB AppDev repos on GitHub"),

    (24, "Phase 5", "Batch 13 - Repos 127 to 136",
     "AB AppDev project, next 10 repos.",
     s(28),  s(29),  "Not Started", "10 AB AppDev repos on GitHub"),

    (25, "Phase 5", "Batch 14 - Repos 137 to 146",
     "AB AppDev project, next 10 repos.",
     s(29),  s(30),  "Not Started", "10 AB AppDev repos on GitHub"),

    (26, "Phase 5", "Batch 15 - Repos 147 to 156",
     "AB AppDev project, next 10 repos.",
     s(30),  s(31),  "Not Started", "10 AB AppDev repos on GitHub"),

    (27, "Phase 5", "Batch 16 - Repos 157 to 166",
     "AB AppDev project, next 10 repos.",
     s(31),  s(32),  "Not Started", "10 AB AppDev repos on GitHub"),

    (28, "Phase 5", "Batch 17 - Repos 167 to 176",
     "AB AppDev project, next 10 repos.",
     s(32),  s(33),  "Not Started", "10 AB AppDev repos on GitHub"),

    (29, "Phase 5", "Batch 18 - Repos 177 to 186",
     "AB AppDev project, next 10 repos.",
     s(33),  s(34),  "Not Started", "10 AB AppDev repos on GitHub"),

    (30, "Phase 5", "Batch 19 - Repos 187 to 196",
     "AB AppDev project, next 10 repos.",
     s(34),  s(35),  "Not Started", "10 AB AppDev repos on GitHub"),

    (31, "Phase 5", "Batch 20 - Repos 197 to 206",
     "AB AppDev project, next 10 repos.",
     s(35),  s(36),  "Not Started", "10 AB AppDev repos on GitHub"),

    (32, "Phase 5", "Batch 21 - Repos 207 to 216",
     "AB AppDev project, next 10 repos.",
     s(36),  s(37),  "Not Started", "10 AB AppDev repos on GitHub"),

    (33, "Phase 5", "Batch 22 - Repos 217 to 226",
     "AB AppDev project, next 10 repos.",
     s(37),  s(38),  "Not Started", "10 AB AppDev repos on GitHub"),

    (34, "Phase 5", "Batch 23 - Repos 227 to 236",
     "AB AppDev project, next 10 repos.",
     s(38),  s(39),  "Not Started", "10 AB AppDev repos on GitHub"),

    (35, "Phase 5", "Batch 24 - Repos 237 to 246",
     "AB AppDev project, next 10 repos.",
     s(38),  s(39),  "Not Started", "10 AB AppDev repos on GitHub"),

    (36, "Phase 5", "Batch 25 - Repos 247 to 256",
     "AB AppDev project, next 10 repos.",
     s(39),  s(40),  "Not Started", "10 AB AppDev repos on GitHub"),

    (37, "Phase 5", "Batch 26 - Repos 257 to 266",
     "AB AppDev project, next 10 repos.",
     s(39),  s(40),  "Not Started", "10 AB AppDev repos on GitHub"),

    (38, "Phase 5", "Batch 27 - Repos 267 to 276",
     "AB AppDev project, next 10 repos.",
     s(40),  s(41),  "Not Started", "10 AB AppDev repos on GitHub"),

    (39, "Phase 5", "Batch 28 - Repos 277 to 286",
     "AB AppDev project, next 10 repos.",
     s(40),  s(41),  "Not Started", "10 AB AppDev repos on GitHub"),

    (40, "Phase 5", "Batch 29 - Repos 287 to 296",
     "AB AppDev project, next 10 repos.",
     s(41),  s(42),  "Not Started", "10 AB AppDev repos on GitHub"),

    (41, "Phase 5", "Batch 30 - Repos 297 to 306",
     "AB AppDev project, next 10 repos.",
     s(41),  s(42),  "Not Started", "10 AB AppDev repos on GitHub"),

    (42, "Phase 5", "Batch 31 - Repos 307 to 309",
     "AB AppDev project, final 3 repos. All 309 active repos are now on GitHub.",
     s(42),  s(42),  "Not Started", "All AB AppDev repos on GitHub"),

    # Phase 6
    (43, "Phase 6", "CI/CD Repo Identification",
     "Identify which repos need a CI/CD pipeline. Document the pipeline type and trigger pattern for each repo before rollout begins.",
     s(39),  s(41),  "Not Started", "CI/CD candidate list approved"),

    (44, "Phase 6", "CI/CD Pipeline Rollout",
     "Apply the standard GitHub Actions workflow YAML to all identified repos using apply_cicd.py. Validate that each pipeline runs successfully on the self-hosted runner.",
     s(39),  s(42),  "Not Started", "Pipelines active, runner jobs passing"),

    (45, "Phase 6", "Full Org Audit",
     "Run post_migration_verify.py across all 309 repos. Every repo must have a project topic, correct default branch, a Ruleset applied, and team access confirmed.",
     s(39),  s(41),  "Not Started", "Audit report with all repos passing"),

    (46, "Phase 6", "Security Review",
     "Confirm no repos are accidentally public, secret scanning is enabled org-wide, push protection is on, and no hardcoded secrets are flagged.",
     s(39),  s(41),  "Not Started", "Security sign-off"),

    (47, "Phase 6", "Consumer URL Updates",
     "Update any submodule references, CI clone URLs, and README badges that still point to ADO. Raise pull requests in each affected repo.",
     s(39),  s(42),  "Not Started", "All consumers updated to GitHub URLs"),

    # Phase 7
    (48, "Phase 7", "ADO Server Archive",
     "Set all ADO repos to read-only. Send communications to all teams: ADO is now archive-only and will be retained for 90 days.",
     s(42),  s(43),  "Not Started", "ADO read-only, comms sent"),

    (49, "Phase 7", "Runbook Handover",
     "Hand over the runbook to the platform team covering: how to add a new repo, apply topics, request access, change a Ruleset, and manage the runner. Includes a 1-hour walkthrough session.",
     s(42),  s(43),  "Not Started", "Runbook delivered and session completed"),

    (50, "Phase 7", "ADO Decommission Planning",
     "Plan the ADO Server decommission for 90 days after archive. Confirm the data retention policy and schedule with the Abank infrastructure team.",
     s(43),  s(43),  "Not Started", "Decommission plan signed off"),
]
write_sheet(
    ws_road,
    ["S.No", "Phase", "Activity", "Description", "Start Date", "End Date", "Status", "Expected Output"],
    roadmap,
    phase_col_idx=1,
)

# ── Sheet 3: Decisions Needed (rows 4-7 removed, renumbered) ──
ws_dec = wb.create_sheet("Decisions Needed")
write_sheet(
    ws_dec,
    ["S.No", "Decision", "Context", "Decision Owner", "Options"],
    [
        (1, "master to main rename",
         "309 repos use master as the default branch. Do we rename to main during migration or keep master?",
         "Project Lead and Dev Teams",
         "Rename to main during the mirror step (recommended) or keep master as-is"),

        (2, "18 empty repos",
         "18 repos have no commits. Do we skip them or create empty shell repos on GitHub?",
         "Repo Owners",
         "Skip the empty repos (recommended) or create empty shells on GitHub"),

        (3, "3 name conflicts",
         "ACH BAI File Generator, BAM, and DW STAGING LOAD LEGACY each exist in more than one project. One copy of each must be renamed before migration.",
         "Repo Owners",
         "Add a project suffix to the less-active copy, for example BAM-standalone"),

        (4, "Self-hosted runner infrastructure",
         "What server or VM will host the org-level self-hosted runner, and who is responsible for maintaining it?",
         "Abank Infra",
         "Dedicated VM (recommended) or reuse an existing build server"),

        (5, "ADO archive retention period",
         "How long should ADO repos remain read-only after cutover before the server is decommissioned?",
         "Abank IT Management",
         "90 days (recommended), 6 months, or permanent"),
    ],
)

# ── Sheet 4: Scripts to Develop ───────────────
ws_scripts = wb.create_sheet("Scripts to Develop")
write_sheet(
    ws_scripts,
    ["S.No", "Phase", "Script Name", "Purpose", "Owner", "Role in Migration"],
    [
        (1, "Phase 1", "git_mirror_migrate.py",
         "Clones each repo from ADO and pushes it to GitHub. Supports dry-run, retry, resume, and master to main rename.",
         "Engineer", "Core migration engine"),

        (2, "Phase 1", "apply_topics_teams.py",
         "Sets GitHub topics and team permissions for each repo based on the workbook. Dry-run and idempotent.",
         "Engineer", "Org hierarchy and access"),

        (3, "Phase 1", "apply_rulesets.py",
         "Creates 3 to 4 Ruleset templates and attaches them to repos by topic. Dry-run supported.",
         "Engineer", "Branch protection"),

        (4, "Phase 1", "post_migration_verify.py",
         "Runs a pass or fail check on each repo: branch count, default branch, topics, team access, and Ruleset.",
         "Engineer", "Quality gate per batch"),

        (5, "Phase 1", "apply_cicd.py",
         "Applies the standard GitHub Actions workflow YAML to identified repos and confirms the pipeline triggers on the self-hosted runner.",
         "Engineer", "CI/CD rollout"),

        (6, "Phase 2", "pilot_runner.sh",
         "Orchestrates the full 5-script sequence on the 2 pilot repos. Captures logs and outputs a summary report.",
         "Engineer", "Pilot orchestration"),
    ],
    phase_col_idx=1,
)

wb.save(OUTPUT)
print(f"Workbook saved to: {OUTPUT}")
